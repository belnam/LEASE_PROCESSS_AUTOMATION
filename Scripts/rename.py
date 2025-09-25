import shutil
import getpass
from datetime import datetime
from pathlib import Path
import fitz  # Make sure you have PyMuPDF installed
import time
import logging
from openpyxl import load_workbook

def processrentals():
    username = getpass.getuser()
    
    # Define your specific paths directly
    output_folder = Path(r'C:\Users\PeaceMuthusi\Documents\RENTALS-PROCESS-AUTOMATION\output_folder')
    input_dir = Path('Downloaded_Scans')  # Make sure this directory exists
    excel_file_path = Path(r'C:\Users\PeaceMuthusi\Documents\RENTALS-PROCESS-AUTOMATION\Rental_ProcessedFilenames\Daily_RentalFileNames.xlsx')
    rename_file_path = Path(r'C:\Users\PeaceMuthusi\Documents\RENTALS-PROCESS-AUTOMATION\Rental_ProcessedFilenames\Renaming_Identifiers.xlsx')
    count_file = Path(r'C:\Users\PeaceMuthusi\Documents\RENTALS-PROCESS-AUTOMATION\count.txt')

    now = datetime.now()
    date_suffix = now.strftime('%B %d %Y')
    time_suffix = now.strftime('%H:%M:%S')

    # Read count from file
    count = 1 if now.day == 1 else int(count_file.read_text().strip())  # Read count from file and strip any whitespace

    pdf_files = [file for file in input_dir.glob('*.pdf') if file.is_file()]

    # Load workbooks
    workbook = load_workbook(excel_file_path)
    rename_workbook = load_workbook(rename_file_path)

    # Read company mappings
    worksheet = rename_workbook["Company_Mappings"]
    company_mapping = {}
    for row in worksheet.iter_rows(values_only=True):
        company_mapping[row[0]] = row[1]

    # Get or create the sheet named "Sheet1"
    if "Sheet1" in workbook.sheetnames:
        output_sheet = workbook["Sheet1"]
    else:
        output_sheet = workbook.create_sheet("Sheet1")

    for pdf_file in pdf_files:
        if not pdf_file.name.startswith('SCAN_'):
            with fitz.open(pdf_file) as pdf_obj:
                text = pdf_obj[0].get_text()

            # Determine company name from text
            company_name = None
            for row in worksheet.iter_rows(values_only=True):
                if row[0].strip('"').lower() in text.lower():  # Strip quotes and match text
                    company_name = row[0]
                    break

            # Determine filename prefix based on company name
            if company_name and company_name in company_mapping:
                prefix = company_mapping[company_name]
            else:
                logging.info(f"Identifier not found for file: {pdf_file}")
                continue

            # Get the original filename
            original_filename = pdf_file.name

            # Construct new filename
            new_filename = f'SCAN_{count}_AMAZON_RENTAL_{prefix}_TOLLS_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')
            new_filename_with_timestamp = new_filename.replace(".pdf", "") + "_" + str(int(time.time())) + ".pdf"

            # Create the full new path for renaming
            new_file_path = output_folder / new_filename_with_timestamp

            try:
                # Check if the original file exists and rename it
                if pdf_file.exists():
                    pdf_file.rename(new_file_path)
                    logging.info(f"File renamed successfully: {new_filename_with_timestamp}")
                else:
                    logging.error(f"File not found: {pdf_file}")
            except Exception as e:
                logging.error(f"Error renaming file {pdf_file}: {str(e)}")
                continue

            # Append details to Excel sheet
            output_sheet.append([original_filename, date_suffix, time_suffix, new_file_path.name.replace('.pdf', '')])
            count += 1

    # Save the Excel workbook with updated records
    workbook.save(excel_file_path)

    # Update the count file
    count_file.write_text(str(count))

    return new_filename_with_timestamp, new_filename, pdf_files

# Run the function
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)  # Set logging level to INFO
    processrentals()

