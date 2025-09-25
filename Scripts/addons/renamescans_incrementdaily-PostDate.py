import shutil
import getpass
from datetime import datetime
from pathlib import Path
import fitz
from openpyxl import load_workbook
from datetime import datetime

username = getpass.getuser()
input_dir = Path('./Downloaded_Scans')
output_dir = Path(f'C:\\Users\\{username}\\Documents\\RENTALS-PROCESS-AUTOMATION\\')
excel_file_path = output_dir / 'Rental_ProcessedFilenames' / 'Daily_RentalFileNames.xlsx'
email_logs_file_path = output_dir / 'Rental_ProcessedFilenames' / 'email_logs.xlsx'
count_file = output_dir / 'count.txt'

now = datetime.now()

count = 1 if now.day == 1 else int(count_file.read_text())  # Read count from file

pdf_files = [file for file in input_dir.glob('*.pdf') if file.is_file()]

workbook = load_workbook(excel_file_path)

# Load email logs workbook
email_logs_workbook = load_workbook(email_logs_file_path)
email_logs_sheet = email_logs_workbook.active

# Read company mappings from the correct Excel sheet
worksheet = workbook["Company_Mappings"]

company_mapping = {}
for row in worksheet.iter_rows(values_only=True):
    company_mapping[row[0]] = row[1]

# Get or create the sheet named "Sheet1"
if "Sheet1" in workbook.sheetnames:
    output_sheet = workbook["Sheet1"]
else:
    output_sheet = workbook.create_sheet("Sheet1")

for pdf_file in pdf_files:
    if not pdf_file.name.startswith('SCAN_'):
        with fitz.open(pdf_file) as pdf_obj:
            text = pdf_obj[0].get_text()

        # Determine company name from text
        company_name = None
        for row in worksheet.iter_rows(values_only=True):
            if row[0].strip('"').lower() in text.lower():  
                company_name = row[0]
                break

        # Determine filename prefix based on company name
        if company_name:
            prefix = company_mapping[company_name]
        else:
            prefix = 'NEW SCAN'

        # Get the original filename before renaming
        original_filename = pdf_file.name

        # Search for original filename in email logs
        received_timestamp = None
        for row in email_logs_sheet.iter_rows(values_only=True):
            if row[0] == original_filename:
                received_timestamp = datetime.strptime(row[1], '%Y-%m-%d %H:%M:%S')  
                break

        # Construct new filename
        if received_timestamp:
            date_suffix = received_timestamp.strftime('%m_%d_%Y')

        new_filename = f'SCAN_{count}_AMAZON_RENTAL_{prefix}_TOLLS_{date_suffix}_(BOT).pdf'

        pdf_file.rename(output_dir / new_filename)

        output_sheet.append([original_filename, date_suffix, new_filename.replace('.pdf', '')])
        count += 1

workbook.save(excel_file_path)
count_file.write_text(str(count))

for entry in output_dir.iterdir():
    try:
        if entry.is_file() and entry.name.lower().endswith('.pdf'):
            shutil.move(entry, input_dir)
    except Exception as e:
        print(f"Error moving file: {entry}")
