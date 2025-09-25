import win32com.client
import os
import pickle
from datetime import datetime
import getpass
import openpyxl


# Set the folder name for pickle files
username = getpass.getuser()
pickle_folder = "PickleFiles"
pdfpath = f"C:\\Users\\{username}\\Documents\\RENTALS-PROCESS-AUTOMATION\\Downloaded_Scans\\"
# Get the path to the pickle files
pickle_file_path_stoughton_account = os.path.join(pickle_folder, "last_processed_timestamp_stoughton_account.pickle")


# Define the date to filter emails
filter_date = datetime(2024, 3, 15) 

rental_accounts_file = f"C:\\Users\\PeaceMuthusi\\Documents\\RENTALS-PROCESS-AUTOMATION\\Scripts\\Rental_Accounts.xlsx"
try:
    wb_rental_accounts = openpyxl.load_workbook(rental_accounts_file)
    ws_rental_accounts = wb_rental_accounts.active
    desired_account_email = ws_rental_accounts['A6'].value  
except Exception as e:
    print("Error loading or reading Rental_Accounts.xlsx:", e)
    desired_account_email = None
 
# Connect to Outlook
outlook = win32com.client.Dispatch("Outlook.Application")
namespace = outlook.GetNamespace("MAPI")

# Get the specific account
accounts = namespace.Accounts
desired_account = None
for account in accounts:
    if account.SmtpAddress == desired_account_email:
        desired_account = account
        break

if desired_account is not None:
    root_folder = namespace.Folders.Item(desired_account.DeliveryStore.DisplayName)

    inbox_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "Inbox":
            inbox_folder = folder
            break
    if inbox_folder is not None:
        inbox_emails = inbox_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'inbox' folder does not exist in the account.")
    inbox_emails.Sort("[ReceivedTime]", False)

    
    last_processed_timestamp_inbox = None
    if os.path.exists(pickle_file_path_stoughton_account):
        try:
            with open(pickle_file_path_stoughton_account, "rb") as file:
                last_processed_timestamp_inbox = pickle.load(file)
        except (EOFError, ValueError):
            pass
    

    if last_processed_timestamp_inbox:
        last_processed_timestamp_inbox = datetime.strptime(last_processed_timestamp_inbox, "%Y-%m-%d %H:%M:%S.%f%z")
    


    # Create or load the log Excel file
    log_file_path = "Rental_ProcessedFilenames/email_logs.xlsx"
    if os.path.exists(log_file_path):
        wb = openpyxl.load_workbook(log_file_path)
    else:
        wb = openpyxl.Workbook()
    ws = wb.active
    

    for email in inbox_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_inbox is None or received_timestamp > last_processed_timestamp_inbox:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_inbox = received_timestamp


    os.makedirs(pickle_folder, exist_ok=True)
    with open(pickle_file_path_stoughton_account, "wb") as file:
        pickle.dump(last_processed_timestamp_inbox.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)

    ws.insert_rows(1)
    ws["A1"] = "Scan Name"
    ws["B1"] = "Received Timestamp"
    ws["C1"] = "Renamed As"
    wb.save(log_file_path)
    wb.close()


else:
    print("Desired account not found.")
