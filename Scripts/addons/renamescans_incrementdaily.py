import shutil,getpass
import util,requests    
from datetime import datetime
from pathlib import Path
import fitz,time ,logging
from openpyxl import load_workbook
from add_to_masterfile import addToMasterfile
from pathlib import Path


def reset_counts(*count_file_paths):
    now = datetime.now()
    current_date = now.date()

    if current_date.day == 1 and not Path(f'reset_done.txt').exists():
        # Reset all count files to 1
        for path in count_file_paths:
            with open(path, 'w') as file:
                file.write('1')
        # Create a file to mark that reset is done for today
        Path(f'reset_done.txt').touch()
    elif current_date.day != 1 and Path(f'reset_done.txt').exists():
        Path(f'reset_done.txt').unlink()



def processrentals():

    output_folder = util.get_uploads_dir()
    input_dir = Path(f'Downloaded_Scans')
    output_dir = Path(f'C:\\Users\\{username}\\Documents\\RENTALS-PROCESS-AUTOMATION\\')
    excel_file_path = output_dir / 'Rental_ProcessedFilenames' / 'Daily_RentalFileNames.xlsx'
    rename_file_path = output_dir / 'Rental_ProcessedFilenames' / 'RenamingIdentifiers.xlsx'
    count_file = output_dir / 'count.txt'

    now = datetime.now()
    date_suffix = now.strftime('%B %d %Y')
    time_suffix = now.strftime('%H:%M:%S')

    count = 1 if now.day == 1 else int(count_file.read_text())  

    pdf_files = [file for file in input_dir.glob('*.pdf') if file.is_file()]

    workbook = load_workbook(excel_file_path)
    rename_workbook = load_workbook(rename_file_path)

    # Read company mappings from the correct Excel sheet
    worksheet = rename_workbook["Company_Mappings"]

    company_mapping = {}
    for row in worksheet.iter_rows(values_only=True):
        company_mapping[row[0]] = row[1]

    # Get or create the sheet named "Sheet1"
    if "Sheet1" in workbook.sheetnames:
        output_sheet = workbook["Sheet1"]
    # else:
    #     output_sheet = workbook.create_sheet("Sheet1")

    for pdf_file in pdf_files:
        if not pdf_file.name.startswith('SCAN_'):
            with fitz.open(pdf_file) as pdf_obj:
                text = pdf_obj[0].get_text()

            # Determine company name from text
            company_name = None
            for row in worksheet.iter_rows(values_only=True):
                if row[0].strip('"').lower() in text.lower():  
                    company_name = row[0]
                    break
            # Determine filename prefix based on company name
            if company_name:
                prefix = company_mapping[company_name]
            else:
                if company_name is None or prefix is None:
                    logging.info(f"Identifier not found for file: {pdf_file}")
                    continue 

            # Get the original filename before renaming
            original_filename = pdf_file.name

            # Construct new filename
            new_filename = f'SCAN_{count}_AMAZON_RENTAL_{prefix}_TOLLS_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')
            new_filename_with_timestamp = new_filename.replace(".pdf", "") + "_" + str(int(time.time())) + ".pdf"

            pdf_file.rename(output_dir / new_filename_with_timestamp)
            new_file_path = output_folder / new_filename_with_timestamp

            logging.info(f"{new_filename_with_timestamp} {new_filename}")
            uploadedToMasterfile, scan_id = addToMasterfile("Amazon",new_filename_with_timestamp, new_filename )
            if uploadedToMasterfile:

                shutil.move(pdf_file, new_file_path)

                #########
                logging.info(new_filename_with_timestamp)
                ###########

                output_sheet.append([original_filename, date_suffix, time_suffix, new_file_path.name.replace('.pdf', '')])
                count += 1

            
            else:
                logging.error(f"Failed To Retrieve Scan ID from Masterfile --  ({uploadedToMasterfile}, {scan_id})")

    workbook.save(excel_file_path)
    count_file.write_text(str(count))
    return new_filename_with_timestamp, new_filename ,pdf_files

processrentals()
    # for entry in output_dir.iterdir():
    #     try:
    #         if entry.is_file() and entry.name.lower().endswith('.pdf'):
    #             shutil.move(entry, input_dir)
    #     except Exception as e:
    #         print(f"Error moving file: {entry}")


