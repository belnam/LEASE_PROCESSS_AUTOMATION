import win32com.client
import os
import pickle
from datetime import datetime
import getpass
import openpyxl


# Set the folder name for pickle files
username = getpass.getuser()
pickle_folder = "PickleFiles"
pdfpath = f"C:\\Users\\{username}\\Documents\\RENTALS-PROCESS-AUTOMATION\\Downloaded_Scans\\"
# Get the path to the pickle files
pickle_file_path_xtra = os.path.join(pickle_folder, "last_processed_timestamp_xtra.pickle")
pickle_file_path_inbox = os.path.join(pickle_folder, "last_processed_timestamp_inbox.pickle")
pickle_file_path_1stadvantage = os.path.join(pickle_folder, "last_processed_timestamp_1stadvantage.pickle")
pickle_file_path_bowman = os.path.join(pickle_folder, "last_processed_timestamp_bowman.pickle")
pickle_file_path_krosstrucking = os.path.join(pickle_folder, "last_processed_timestamp_krosstrucking.pickle")
pickle_file_path_milestone = os.path.join(pickle_folder, "last_processed_timestamp_milestone.pickle")
pickle_file_path_premier = os.path.join(pickle_folder, "last_processed_timestamp_premier.pickle")
pickle_file_path_stoughton = os.path.join(pickle_folder, "last_processed_timestamp_stoughton.pickle")


# Define the date to filter emails
filter_date = datetime(2024, 3, 15) 

rental_accounts_file = f"C:\\Users\\PeaceMuthusi\\Documents\\RENTALS-PROCESS-AUTOMATION\\Scripts\\Rental_Accounts.xlsx"
try:
    wb_rental_accounts = openpyxl.load_workbook(rental_accounts_file)
    ws_rental_accounts = wb_rental_accounts.active
    # Assuming the email address is in the first row of the column named "ACCOUNTS"
    desired_account_email = ws_rental_accounts['A2'].value  
except Exception as e:
    print("Error loading or reading Rental_Accounts.xlsx:", e)
    desired_account_email = None
 
# Connect to Outlook
outlook = win32com.client.Dispatch("Outlook.Application")
namespace = outlook.GetNamespace("MAPI")

# Get the specific account
accounts = namespace.Accounts
desired_account = None
for account in accounts:
    if account.SmtpAddress.lower() == desired_account_email:
        desired_account = account
        break

if desired_account is not None:
    root_folder = namespace.Folders.Item(desired_account.DeliveryStore.DisplayName)

    xtra_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "XTRA":
            xtra_folder = folder
            break
    if xtra_folder is not None:
        xtra_emails = xtra_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'XTRA' folder does not exist in the account.")
    xtra_emails.Sort("[ReceivedTime]", False)

    inbox_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "Inbox":
            inbox_folder = folder
            break
    if inbox_folder is not None:
        inbox_emails = inbox_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'inbox' folder does not exist in the account.")
    inbox_emails.Sort("[ReceivedTime]", False)

    firstadvantage_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "1ST ADVANTAGE":
            firstadvantage_folder = folder
            break
    if firstadvantage_folder is not None:
        firstadvantage_emails = firstadvantage_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'firstadvantage' folder does not exist in the account.")
    firstadvantage_emails.Sort("[ReceivedTime]", False)

    bowman_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "Bowman Rentals":
            bowman_folder = folder
            break
    if bowman_folder is not None:
        bowman_emails = bowman_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'bowman' folder does not exist in the account.")
    bowman_emails.Sort("[ReceivedTime]", False)
    krosstrucking_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "KROSS TRUCKING":
            krosstrucking_folder = folder
            break
    if krosstrucking_folder is not None:
        krosstrucking_emails = krosstrucking_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'krosstrucking' folder does not exist in the account.")
    krosstrucking_emails.Sort("[ReceivedTime]", False)
    milestone_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "Milestone":
            milestone_folder = folder
            break
    if milestone_folder is not None:
        milestone_emails = milestone_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'milestone' folder does not exist in the account.")
    milestone_emails.Sort("[ReceivedTime]", False)

    premier_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "Premier":
            premier_folder = folder
            break
    if premier_folder is not None:
        premier_emails = premier_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'Premier' folder does not exist in the account.")
    premier_emails.Sort("[ReceivedTime]", False)

    stoughton_folder = None
    for folder in root_folder.Folders:
        if folder.Name == "Stoughton Rental":
            stoughton_folder = folder
            break
    if stoughton_folder is not None:
        stoughton_emails = stoughton_folder.Items.Restrict("[ReceivedTime] > '{}'".format(filter_date.strftime("%m/%d/%Y")))
    else:
        print("The 'stoughton' folder does not exist in the account.")
    stoughton_emails.Sort("[ReceivedTime]", False)


    last_processed_timestamp_xtra = None
    if os.path.exists(pickle_file_path_xtra):
        try:
            with open(pickle_file_path_xtra, "rb") as file:
                last_processed_timestamp_xtra = pickle.load(file)
        except (EOFError, ValueError):
            pass
    last_processed_timestamp_inbox = None
    if os.path.exists(pickle_file_path_inbox):
        try:
            with open(pickle_file_path_inbox, "rb") as file:
                last_processed_timestamp_inbox = pickle.load(file)
        except (EOFError, ValueError):
            pass
    last_processed_timestamp_1stadvantage = None
    if os.path.exists(pickle_file_path_1stadvantage):
        try:
            with open(pickle_file_path_1stadvantage, "rb") as file:
                last_processed_timestamp_1stadvantage = pickle.load(file)
        except (EOFError, ValueError):
            pass
    last_processed_timestamp_bowman = None
    if os.path.exists(pickle_file_path_bowman):
        try:
            with open(pickle_file_path_bowman, "rb") as file:
                last_processed_timestamp_bowman = pickle.load(file)
        except (EOFError, ValueError):
            pass
    last_processed_timestamp_krosstrucking = None
    if os.path.exists(pickle_file_path_krosstrucking):
        try:
            with open(pickle_file_path_krosstrucking, "rb") as file:
                last_processed_timestamp_krosstrucking = pickle.load(file)
        except (EOFError, ValueError):
            pass
    last_processed_timestamp_milestone = None
    if os.path.exists(pickle_file_path_milestone):
        try:
            with open(pickle_file_path_milestone, "rb") as file:
                last_processed_timestamp_milestone = pickle.load(file)
        except (EOFError, ValueError):
            pass
    last_processed_timestamp_premier = None
    if os.path.exists(pickle_file_path_premier):
        try:
            with open(pickle_file_path_premier, "rb") as file:
                last_processed_timestamp_premier = pickle.load(file)
        except (EOFError, ValueError):
            pass
    last_processed_timestamp_stoughton = None
    if os.path.exists(pickle_file_path_stoughton):
        try:
            with open(pickle_file_path_stoughton, "rb") as file:
                last_processed_timestamp_stoughton = pickle.load(file)
        except (EOFError, ValueError):
            pass

    if last_processed_timestamp_xtra:
        last_processed_timestamp_xtra = datetime.strptime(last_processed_timestamp_xtra, "%Y-%m-%d %H:%M:%S.%f%z")
    if last_processed_timestamp_inbox:
        last_processed_timestamp_inbox = datetime.strptime(last_processed_timestamp_inbox, "%Y-%m-%d %H:%M:%S.%f%z")
    if last_processed_timestamp_1stadvantage:
        last_processed_timestamp_1stadvantage = datetime.strptime(last_processed_timestamp_1stadvantage, "%Y-%m-%d %H:%M:%S.%f%z")
    if last_processed_timestamp_bowman:
        last_processed_timestamp_bowman = datetime.strptime(last_processed_timestamp_bowman, "%Y-%m-%d %H:%M:%S.%f%z")
    if last_processed_timestamp_krosstrucking:
        last_processed_timestamp_krosstrucking = datetime.strptime(last_processed_timestamp_krosstrucking, "%Y-%m-%d %H:%M:%S.%f%z")
    if last_processed_timestamp_milestone:
        last_processed_timestamp_milestone = datetime.strptime(last_processed_timestamp_milestone, "%Y-%m-%d %H:%M:%S.%f%z")
    if last_processed_timestamp_premier:
        last_processed_timestamp_premier = datetime.strptime(last_processed_timestamp_premier, "%Y-%m-%d %H:%M:%S.%f%z")
    if last_processed_timestamp_stoughton:
        last_processed_timestamp_stoughton = datetime.strptime(last_processed_timestamp_stoughton, "%Y-%m-%d %H:%M:%S.%f%z")


    # Create or load the log Excel file
    log_file_path = "Rental_ProcessedFilenames/email_logs.xlsx"
    if os.path.exists(log_file_path):
        wb = openpyxl.load_workbook(log_file_path)
    else:
        wb = openpyxl.Workbook()
    ws = wb.active
    
    for email in xtra_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_xtra is None or received_timestamp > last_processed_timestamp_xtra:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_xtra = received_timestamp

    for email in inbox_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_inbox is None or received_timestamp > last_processed_timestamp_inbox:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_inbox = received_timestamp

    for email in firstadvantage_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_1stadvantage is None or received_timestamp > last_processed_timestamp_1stadvantage:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_1stadvantage = received_timestamp

    for email in bowman_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_bowman is None or received_timestamp > last_processed_timestamp_bowman:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_bowman = received_timestamp

    for email in krosstrucking_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_krosstrucking is None or received_timestamp > last_processed_timestamp_krosstrucking:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_krosstrucking = received_timestamp

    for email in milestone_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_milestone is None or received_timestamp > last_processed_timestamp_milestone:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_milestone = received_timestamp

    for email in premier_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_premier is None or received_timestamp > last_processed_timestamp_premier:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_premier = received_timestamp

    for email in stoughton_emails:
        received_timestamp = email.ReceivedTime
        if last_processed_timestamp_stoughton is None or received_timestamp > last_processed_timestamp_stoughton:
            attachments = email.Attachments
            count = attachments.Count
            for j in range(1, count + 1):
                attachment = attachments.Item(j)
                file_extension = os.path.splitext(attachment.FileName)[1].lower()
                if file_extension == ".pdf":
                    save_path = os.path.join(pdfpath, attachment.FileName)
                    attachment.SaveAsFile(save_path)
                    received_timestamp_str = received_timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([attachment.FileName, received_timestamp_str])
            last_processed_timestamp_stoughton = received_timestamp

    os.makedirs(pickle_folder, exist_ok=True)
    with open(pickle_file_path_xtra, "wb") as file:
        pickle.dump(last_processed_timestamp_xtra.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)
    with open(pickle_file_path_inbox, "wb") as file:
        pickle.dump(last_processed_timestamp_inbox.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)
    with open(pickle_file_path_1stadvantage, "wb") as file:
        pickle.dump(last_processed_timestamp_1stadvantage.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)
    with open(pickle_file_path_bowman, "wb") as file:
        pickle.dump(last_processed_timestamp_bowman.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)
    with open(pickle_file_path_krosstrucking, "wb") as file:
        pickle.dump(last_processed_timestamp_krosstrucking.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)
    with open(pickle_file_path_milestone, "wb") as file:
        pickle.dump(last_processed_timestamp_milestone.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)
    with open(pickle_file_path_premier, "wb") as file:
        pickle.dump(last_processed_timestamp_premier.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)
    with open(pickle_file_path_stoughton, "wb") as file:
        pickle.dump(last_processed_timestamp_stoughton.strftime("%Y-%m-%d %H:%M:%S.%f%z"), file)

    ws.insert_rows(1)
    ws["A1"] = "Scan Name"
    ws["B1"] = "Received Timestamp"
    ws["C1"] = "Renamed As"
    wb.save(log_file_path)
    wb.close()


else:
    print("Desired account not found.")
