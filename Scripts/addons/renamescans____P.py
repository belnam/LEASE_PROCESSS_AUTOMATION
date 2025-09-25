import shutil
from datetime import datetime
from pathlib import Path
import fitz, time
from openpyxl import load_workbook
import requests
from add_to_masterfile import addToMasterfile
import logging, util
from convertPDF import convert_single_scanned_pdf_to_searchable

def reset_counts(*count_file_paths):
    now = datetime.now()
    current_date = now.date()

    if current_date.day == 1 and not Path(f'{util.get_working_dir()}count_files/reset_done.txt').exists():
        # Reset all count files to 1
        for path in count_file_paths:
            with open(path, 'w') as file:
                file.write('1')
        # Create a file to mark that reset is done for today
        Path(f'{util.get_working_dir()}count_files/reset_done.txt').touch()
    elif current_date.day != 1 and Path(f'{util.get_working_dir()}count_files/reset_done.txt').exists():
        Path(f'{util.get_working_dir()}count_files/reset_done.txt').unlink()

def process_amazon_pdfs():
    input_dir = util.get_amazon_dir()
    output_folder = util.get_uploads_dir()
    excel_file_path = f"{util.get_working_dir()}Renaming_Identifiers.xlsx"
    count_file_path = f'{util.get_working_dir()}count_files/count.txt'
    count_ntta_file_path = f'{util.get_working_dir()}count_files/count_ntta.txt'
    count_njta_file_path = f'{util.get_working_dir()}count_files/count_njta.txt'
    count_citation_file_path = f'{util.get_working_dir()}count_files/count_citation.txt'

    reset_counts(count_file_path, count_ntta_file_path, count_njta_file_path)

    now = datetime.now()
    date_suffix = now.strftime('%m-%d-%Y')
    time_suffix = now.strftime('%H:%M:%S')

    new_filename_with_timestamp = None
    new_filename = None

    with open(count_file_path, 'r') as file:
        last_count = int(file.read())

    count = last_count
    pdf_files = [file for file in input_dir.glob('*.pdf') if file.is_file()]

    workbook = load_workbook(excel_file_path)
    worksheet = workbook["Company_Mappings"]

    company_mapping = {}
    for row in worksheet.iter_rows(values_only=True):
        company_mapping[row[0]] = (row[1], row[2])

    if "Sheet1" in workbook.sheetnames:
        output_sheet = workbook["Sheet1"]
    else:
        output_sheet = workbook.create_sheet("Sheet1")
    new_file_path = None

    for pdf_file in pdf_files:
        convert_single_scanned_pdf_to_searchable(pdf_file)
        if not pdf_file.name.startswith('SCAN_'):
            with fitz.open(pdf_file) as pdf_obj:
                text = ""
                for page in pdf_obj:
                    text += page.get_text()
            company_name = None
            prefix = None
            for row in worksheet.iter_rows(values_only=True):
                if row[0].strip('"').lower() in text.lower():
                    company_name = row[0]
                    prefix, value = company_mapping[company_name]
                    break
             # skip file if identifier is not found
            if company_name is None or prefix is None:
                logging.info(f"Identifier not found for file: {pdf_file}")
                continue 
            

            if prefix:
                original_filename = pdf_file.name
                if value == 'njta':
                    with open(count_njta_file_path, 'r') as file:
                        count = int(file.read())
                    new_filename = f'SCAN_{count}_{prefix}_AMAZON_NJTA_TOLLS_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')
                    new_filename_with_timestamp = new_filename.replace(".pdf", "") + "_" + str(int(time.time())) + ".pdf"

                    with open(count_njta_file_path, 'w') as file:
                        file.write(str(count + 1))
                elif value == 'ntta':
                    with open(count_ntta_file_path, 'r') as file:
                        count = int(file.read())
                    new_filename = f'SCAN_{count}_{prefix}_AMAZON_NTTA_TOLLS_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')
                    new_filename_with_timestamp = new_filename.replace(".pdf", "") + "_" + str(int(time.time())) + ".pdf"

                    with open(count_ntta_file_path, 'w') as file:
                        file.write(str(count + 1))
                elif value == 'citation':
                    with open(count_citation_file_path, 'r') as file:
                        count = int(file.read())
                    new_filename = f'SCAN_{count}_AMAZON_CITATION_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')
                    new_filename_with_timestamp = new_filename.replace(".pdf", "") + "_" + str(int(time.time())) + ".pdf"

                    with open(count_citation_file_path, 'w') as file:
                        file.write(str(count + 1))

                else:
                    with open(count_file_path, 'r') as file:
                        count = int(file.read())
                    new_filename = f'SCAN_{count}_{prefix}_AMAZON_TOLLS_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')
                    new_filename_with_timestamp = new_filename.replace(".pdf", "") + "_" + str(int(time.time())) + ".pdf"

                    with open(count_file_path, 'w') as file:
                        file.write(str(count + 1))
                new_file_path = output_folder / new_filename_with_timestamp

                while new_file_path.exists():
                    count += 1
                    if value == 'njta':
                        new_filename = f'SCAN_{count}_{prefix}_AMAZON_NJTA_TOLLS_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')

                    elif value == 'ntta':
                        new_filename = f'SCAN_{count}_{prefix}_AMAZON_NTTA_TOLLS_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')

                    elif value == 'citation':
                        new_filename = f'SCAN_{count}_AMAZON_CITATION-{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')

                    else:
                        new_filename = f'SCAN_{count}_{prefix}_AMAZON_TOLLS_{date_suffix}_(BOT).pdf'.lower().replace(' ', '_')

                        new_filename_with_timestamp = new_filename.replace(".pdf", "") + "_" + str(int(time.time())) + ".pdf"
                    new_file_path = output_folder / new_filename_with_timestamp

                logging.info(f"{new_filename_with_timestamp} {new_filename}")
                uploadedToMasterfile, scan_id = addToMasterfile("Amazon",new_filename_with_timestamp, new_filename )
                if uploadedToMasterfile:

                    shutil.move(pdf_file, new_file_path)

                    #########
                    logging.info(new_filename_with_timestamp)
                    ###########

                    output_sheet.append([original_filename, date_suffix, time_suffix, new_file_path.name.replace('.pdf', '')])
                    count += 1

                    ## Send File to Vincent
                    url = f"{util.getMLUrl()}/upload_srt?scan_id={scan_id}"
                 
                    files = {'srt_file': (new_filename_with_timestamp, open(new_file_path, 'rb'), 'application/pdf', {'Expires': '0'})}
                    logging.info(files)
                    res = requests.post(url, files=files)
                    logging.info(res.headers)
                    logging.info(res.content)
                else:
                    logging.error(f"Failed To Retrieve Scan ID from Masterfile --  ({uploadedToMasterfile}, {scan_id})")

    workbook.save(excel_file_path)
    workbook.close()

    return new_filename_with_timestamp, new_filename ,pdf_files

# process_amazon_pdfs()


